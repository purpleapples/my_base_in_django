import copy


def set_int_to_excel_column(index):
    quotient = index // 26
    remainder = index % 26
    column = str(chr(65 + remainder))
    if quotient > 1:
        column = str(chr(65 + quotient)) + column
    return column

class CellCopier:
    """
        기존의 cell 의 양식 및 값에 대한 복사 붙여넣기 기능 및 삽입과 동시에 기존 양식 보존 기능
        작성 필요 기능
        push_and_insert(origin_coordinate, target_coordinate, keep_value=False)
        copy_width_and_height(origin_coordinate, target_coordinate)
    """

    def __init__(self, wb, sheet_name):
        self.sheet = wb[sheet_name]
        self.area_dict = dict()

    def copy_width_and_height(self, origin_coordinate, target_coordinate):
        pass

    def push_and_insert(self, key, move_base='row', start_row_and_col=(0,0) ,count=0, keep_value=True):
        coordinate = self.area_dict[key]['coordinate']
        copy_target_rows = self.area_dict[key]['rows']
        """ push rows and keep their format and insert a bunch of specific area there
        :param str key : key for area_dict
        :param str move_base: where copy
        :param tuple[int, int] start_row_and_col: start coordinate
        :param int count: how many area will be copy and insert into sheet
        :param keep_value: keep cell value when cell is not mergedCell
        :return: 
        """
        start_row = start_row_and_col[0]
        start_col = start_row_and_col[1]
        sheet_end_row = len(list(self.sheet.rows))
        rows = copy.deepcopy(list(self.sheet.rows)[start_row: sheet_end_row])
        row_list = [row for row in range(start_row, sheet_end_row)]
        merge_range_str_list = list()

        # key : start cell coordinate, value = row range, col range
        import re
        regex_str = '[\d]{1,}'
        def set_coordinate_from_merge_cell(coordinate):
            if isinstance(coordinate, str):
                coordinate_list = coordinate.split(':')
                start = coordinate_list[0]
                end = coordinate_list[1]
                rows = re.findall(regex_str, coordinate)
                start_row = rows[0]
                end_row = rows[1]
                start_column = start[:start.index(start_row)]
                end_column =  end[:end.index(end_row)]
                start_column_index = 0
                end_column_index = 0
                for index, c in enumerate(start_column[::-1]):
                    column = int(ord(c)) - 64
                    if index >0:
                        column *= index * 26
                    start_column_index += column
                # column to int
                start_row = int(start_row)
                end_row = int(end_row)
                for index, c in enumerate(end_column[::-1]):
                    column = int(ord(c)) - 64
                    if index > 0:
                        column *= index * 26
                    end_column_index += column
                return (end_row - start_row , end_column_index - start_column_index)
            elif isinstance(coordinate, tuple):
                pass
            return ''

        merge_range_dict = {str(cell).split(':')[0]:set_coordinate_from_merge_cell(str(cell)) for cell in self.sheet.merged_cells}
        if move_base =='row':
            interval = coordinate['end_row'] - coordinate['start_row']
            insert_end_row = start_row + interval * count
            row_dimensions = copy.deepcopy([self.sheet.row_dimensions[row] for row in row_list])
            ############################## move below rows before insert ###############################################
            ####### unmerge cells
            merge_row = insert_end_row - start_row -1
            for cell in list(self.sheet.merged_cells):
                if cell.min_row > start_row:
                    # 문자를 변경 후 range string 저장시킨다.
                    range_string = str(cell)
                    self.sheet.unmerge_cells(range_string=range_string)
                    range_string = range_string.replace(str(cell.min_row), str(cell.min_row + merge_row))
                    range_string = range_string.replace(str(cell.max_row), str(cell.max_row + merge_row))
                    merge_range_str_list.append(range_string)

            ####### remove style, value in insertion area
            for row in list(self.sheet.rows)[start_row: sheet_end_row]:
                for cell in row:
                    cell.value = None
                    # cell._style = None

            ####### copy below cells style and values
            for dim, row, copy_row in zip(row_dimensions, rows, range(insert_end_row, insert_end_row + len(row_dimensions))):
                row_index = dim.r + insert_end_row
                self.sheet.row_dimensions[row_index].height = dim.height

                for source_cell in row:
                    copy_cell = self.sheet.cell(row=copy_row, column=source_cell.column)

                    if not isinstance(source_cell, openpyxl.cell.cell.MergedCell):
                        value = source_cell._value
                        if value is not None:
                            copy_cell._value = value
                            copy_cell.data_type = source_cell.data_type
                    if source_cell.has_style:
                        copy_cell._style = copy.copy(source_cell._style)

                    if source_cell.comment:
                        copy_cell.comment = copy.copy(source_cell.comment)

            for merge_range_str in merge_range_str_list:
                self.sheet.merge_cells(range_string=merge_range_str)
            ############################# insert rows and copy cell style ##############################################

            for row_index in range(0, count):
                for add_row, source_row in enumerate(copy_target_rows):
                    row = start_row + row_index * len(copy_target_rows) + add_row

                    for source_cell in source_row:

                        column = source_cell.column + start_col
                        copy_cell = self.sheet.cell(row=row, column=column)

                        if source_cell.coordinate in merge_range_dict.keys():
                            move_coordinate = merge_range_dict[source_cell.coordinate]
                            self.sheet.merge_cells(start_row=row, end_row=row + move_coordinate[0],
                                                   start_column=column, end_column=column + move_coordinate[1])

                        if not keep_value:
                            if not isinstance(copy_cell, openpyxl.cell.cell.MergedCell):
                                copy_cell.value = None
                        else:
                            if not isinstance(copy_cell, openpyxl.cell.cell.MergedCell):
                                copy_cell.value = source_cell.value
                        if source_cell.has_style:
                            copy_cell._style = copy.copy(source_cell._style)
                        if source_cell.comment:
                            copy_cell.comment = copy.copy(source_cell.comment)
                        copy_cell.border = copy.copy(source_cell.border)

        else:
            move_end_col = coordinate['end_column'] - coordinate['start_column']

    def set_area(self, key, coordinate=dict(start_row=1, end_row=1, start_column=1, end_column=1), keep_align=True, keep_value=False):
        # not include end number
        row_dimensions = [self.sheet.row_dimensions[index] for index in
                          range(coordinate['start_row'],coordinate['end_row'])]

        # index 가 26을 넘을 경우 26 의 나머지만 챙기고 26의 목을 가지고는 앞단 계산을 실시한다.
        def set_int_to_excel_column(index):
            quotient = index // 26
            remainder = index % 26
            column = str(chr(65+remainder))
            column = str(chr(65+remainder))
            if quotient > 1:
                column = str(chr(65+quotient))+ column
            return column

        column_list = [set_int_to_excel_column(column)  for column in  range(coordinate['start_column'], coordinate['end_column'])]
        column_dimensions = { column:self.sheet.column_dimensions[column] for column in column_list}

        # for index, row in enumerate(self.sheet.rows):

        self.area_dict[key] = dict(
            row_dimensions = row_dimensions,
            column_dimensions = column_dimensions,
            rows = list(self.sheet.rows)[coordinate['start_row']:coordinate['end_row']],
            coordinate = coordinate
        )
        # pass

    def get_area(self, key):
        return self.area_dict[key] if key in self.area_dict.keys() else None

    def change_rows(self, key, sheet, start_row):
        area_info = self.area_dict[key]
        target_rows = list(sheet.rows)[start_row:len(area_info['row_dimensions'])]

        for index, row_dimension in enumerate(area_info['row_dimensions']):
            target_rows[index].height = row_dimension.height
        for column, column_dimension in area_info['column_dimensions'].items():
            sheet.column_dimensions[column].width = column_dimension.width

        for row in range(0, len(area_info['row_dimensions'])):
            for column in area_info['column_dimensions'].keys():
                if len(column) == 2 :
                    column = (int(ord(column[0])) - 65) * 26 + int(ord(column[1])) - 64
                else:
                    column = ord(column) - 64

                sheet.cell(row +start_row, column).fill = self.sheet.cell(row, column).fill
                sheet.cell(row + start_row, column).border = self.sheet.cell(row, column).border
                sheet.cell(row + start_row, column).alignment = self.sheet.cell(row, column).aligment
                # sheet.cell(row + start_row, column).fill = self.sheet.cell(row, column).fill
                # sheet.cell(row + start_row, column).fill = self.sheet.cell(row, column).fill

        # start_row 아래 부분을 전부 옮겨야 한다.

    def copy_cell_area(self, move_row=0, move_col=0, target=None):

        if target is not None:
            if move_row != 0:
                self.set_row_dimension(move_row, target)
            if move_col != 0:
                self.set_column_dimension(move_col, target)
        else:
            if move_row != 0:
                self.set_row_dimension(move_row)
            if move_col != 0:
                self.set_column_dimension(move_col)

    def copy_cell_content(self ,move_row=0, move_col=0, target=None):
        import openpyxl
        from copy import copy
        if target is not None:
            for (row, col), source_cell in self.origin_cell_info.items():
                target_cell = target.cell(column=col + move_col, row=row + move_row)

                if type(target_cell) != openpyxl.cell.cell.MergedCell:
                    target_cell._value = source_cell._value
                    target_cell.data_type = source_cell.data_type

                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)

                if source_cell.hyperlink:
                    target_cell._hyperlink = copy(source_cell.hyperlink)

                if source_cell.comment:
                    target_cell.comment = copy(source_cell.comment)
        else:
            for (row, col), source_cell in self.origin_cell_info.items():
                target_cell = self.sheet.cell(column=col+move_col, row=row+move_row)

                if type(target_cell) != openpyxl.cell.cell.MergedCell:
                    target_cell._value = source_cell._value
                    target_cell.data_type = source_cell.data_type

                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)

                if source_cell.hyperlink:
                    target_cell._hyperlink = copy(source_cell.hyperlink)

                if source_cell.comment:
                    target_cell.comment = copy(source_cell.comment)
